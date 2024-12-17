#!/usr/bin/python

import sys
import json
import argparse
import yaml
import pandas as pd
import os
from pydantic import BaseModel
from dataclasses import dataclass
from openpyxl import Workbook
import tiktoken
from openai import OpenAI, Model

def encoding_getter(encoding_type: str):
    """
    Returns the appropriate encoding based on the given encoding type (either an encoding string or a model name).
    """
    if "k_base" in encoding_type:
        return tiktoken.get_encoding(encoding_type)
    else:
        return tiktoken.encoding_for_model(encoding_type)

def tokenizer(string: str, encoding_type: str) -> list:
    """
    Returns the tokens in a text string using the specified encoding.
    """
    encoding = encoding_getter(encoding_type)
    tokens = encoding.encode(string)
    return tokens

def token_counter(string: str, encoding_type: str) -> int:
    """
    Returns the number of tokens in a text string using the specified encoding.
    """
    num_tokens = len(tokenizer(string, encoding_type))
    return num_tokens


# This section is for flags
DEBUG = None
def select_model(name):
    global MODEL
    MODEL = name
    global CLIENT
    # Do not know why, but "global" fixes it
    CLIENT = OpenAI()

def select_model_from_nickname(nickname):
    if nickname in MODEL_NICKNAMES:
        model = MODEL_NICKNAMES[nickname]
        select_model(model)
        return True
    else:
        print("Nickname {} not found!".format(nickname))
        print("Crashing since we don't know what AI Model to use.")
        print("Select one of these model names using `--model MODEL_NAME`:")
        for alias, model in MODEL_NICKNAMES:
            print("\t{}: Uses {}".format(alias,model))
        return False

class DictObject(object):
    def __init__(self, dict_):
        self.__dict__.update(dict_)

    @classmethod
    def from_dict(cls, d):
        return json.loads(json.dumps(d), object_hook=DictObject)


def prompt(system_prompt, question, response_model=None):
    args = {
        'model': MODEL,
        'messages': [
            {
                'role': 'system',
                'content': system_prompt
            },
            {
                'role': 'user',
                'content': question
            }
        ]
    }
    if response_model:
        # args['temperature'] = 0
        args['response_format'] = response_model 
    response = None
    completion = None

    if response_model:
        completion = CLIENT.beta.chat.completions.parse
    else:
        completion = CLIENT.chat.completions.create
    completion = completion(**args)
    response = completion.choices[0]

    if response_model:
        response = response.message.parsed
    else:
        response = response.message.content
    
    """
    if DEBUG:
        input_length = token_counter(system_prompt, MODEL) + token_counter(question, MODEL)
        input_cost = (input_length / 1000) * 0.0025
        output_length = token_counter(str(response), MODEL)
        output_cost = (output_length / 1000) * 0.01
        print("Model", MODEL, "will require", input_length, "input tokens and",  output_length," output tokens to process this data.")
        print("That is", str(input_length/1280) + "%", "of input possible and", str(output_length/1280) + "%", "of output possible.")
        print("Cost is", "$" + str(input_cost), "for input and", "$" + str(output_cost), "for output or ", "$" + str(input_cost + output_cost), "total.")
    """

    return response

def save_dicts_to_csv(data, output_filename):
    """
    This function is also AI Generated.

    Save a dictionary of dictionaries to a single CSV file, with an added column
    indicating the original sheet (dictionary) name for each row.
    
    Parameters:
    - data (dict of dict): A dictionary where each key represents a sheet name, and each
                           value is a dictionary representing data for that sheet.
    - output_filename (str): The path and name of the output CSV file.
    """
    # Create an empty list to hold DataFrames for each sheet
    all_dataframes = []
    
    for sheet_name, sheet_data in data.items():
        # Convert the current dictionary to a DataFrame
        df = pd.DataFrame(sheet_data)
        # Add a column for the sheet name
        df['Category'] = sheet_name
        # Append to the list of DataFrames
        all_dataframes.append(df)
    
    # Concatenate all DataFrames into a single DataFrame
    final_df = pd.concat(all_dataframes, ignore_index=True)
    
    # Write the final DataFrame to a CSV file
    final_df.to_csv("./data/" + output_filename, index=False)
    print(f"Data successfully saved to {output_filename}")

def write_to_json(data, filename):
    with open(get_filename(filename, 'json'), 'w') as f:
        json.dump(data, f, indent=4)
    print("Data successfully written to ", get_filename(filename, 'json'))

def write_to_excel(data, output_filename):
    """
    This function is mostly AI Generated.

    Save a dictionary of dictionaries to an Excel file, with each subdictionary
    in a separate sheet.
    
    Parameters:
    - data (dict of dict): A dictionary where each key represents the sheet name,
                           and each value is a dictionary representing data for that sheet.
    - output_filename (str): The path and name of the output Excel file.
    """
    with pd.ExcelWriter(output_filename) as writer:
        for sheet_name, sheet_data in data.items():
            # Convert each subdictionary to a DataFrame
            df = pd.DataFrame(sheet_data)
            # Write the DataFrame to a specific sheet in the Excel file
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    print(f"Data successfully saved to {output_filename}")


def load_config():
    with open('./config.yaml', 'r') as f:
        config = yaml.safe_load(f)
    return config

CONFIG = load_config()

PROMPTS = CONFIG['PROMPTS']
MODEL_NICKNAMES = CONFIG['MODEL_NICKNAMES']

def prompt_sort(curlist):
    return PROMPTS["SORT"] + ", ".join(curlist)

    print("\n\n\n[\n")
    for i in l:
        print("\t", i, "\n")
    print("]\n\n\n")

def dpr(*args):
    if DEBUG:
        print(*args)

def convert(excel_file_path):
    """
    Converts all sheets in an Excel file to separate CSV files.
    
    Parameters:
    - excel_file_path: str, path to the Excel file
    - output_dir: str, directory where the CSV files should be saved
    
    Each CSV file will have the same name as the sheet in the Excel file.
    """
    
    output_dir = excel_file_path.split(".")[0]

    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Load all sheets into a dictionary
    excel_data = pd.read_excel(excel_file_path, engine='openpyxl', sheet_name=None)  # sheet_name=None reads all sheets
    
    # Iterate over each sheet and save as CSV
    for sheet_name, data in excel_data.items():
        # Define the output CSV file path
        csv_file_path = os.path.join(output_dir, f"{sheet_name}.csv")
        dpr(data, "\n") 
        # Save each sheet's data to a CSV file
        dpr(csv_file_path)
        data.to_csv(csv_file_path, index=None)
        dpr(f"Saved sheet '{sheet_name}' to {csv_file_path}")


def summarize_req(r):
    data = json.dumps(r)
    dpr(data)
    res = prompt(PROMPTS["SUMMARIZE"], data)
    dpr(res, '\n')
    return res

class Topic(BaseModel):
    name: str
    description: str

def sort_req(summary, commonalities):
    dpr(summary)
    mprompt = prompt_sort(commonalities) # Generates Prompt
    res = prompt(mprompt, summary, Topic)
    dpr(res, "\n")
    return res

def get_filename(f, ext):
    return "data/{}.{}".format(f, ext)

class Reader:
    def __init__(self, filename: str):
        self.filename = filename

    # Get all of the data
    def __read__(self):
        self.data = []

    def __iter__(self):
        self.index = -1
        dpr("***{}***\n".format(self.filename))
        self.__read__()
        return self

    def __next__(self):
        self.index += 1
        if self.index == len(self.data):
            raise StopIteration
        dpr("{} of {}".format(self.index + 1, len(self.data)))
        return self.data[self.index]


def read_csv(f):
    return pd.read_csv(get_filename(f, 'csv')).to_dict('records')

def read_json(f):
    return [r[0] for r in pd.read_json(get_filename(f, "json")).values]

# Iterates through a CSV File
class CSVReader(Reader):
    def __read__(self):
        self.data = read_csv(self.filename)

class JSONReader(Reader):
    def __read__(self):
        # For some reason, it returns obj[0] insetead of obj for each, and I don't know why
        self.data = read_json(self.filename)

class SortReader(Reader):
    def __read__(self):
        self.data = [
            {
                'request': d[0],
                'summary': d[1]
            }
            for d in zip(read_csv(self.filename), read_json(self.filename))
        ]

def get_target_list(ts):
    # ts = "ALL" or base of filename
    target_list = [ts]
    if ts == "ALL":
        target_list = list(set([f.split(".")[0] for f in os.listdir("data")]))
    return target_list

# This wraps any function with the get_target_list function
# get_target_list takes either a list of file root or "ALL" and turns that into a list of filesnames
def targeter(fn):
    def wrap(t):
        fn(get_target_list(t))
    return wrap

def summarize_file(f):
    summaries = []
    for data in CSVReader(f):
        summaries.append(summarize_req(data))
        write_to_json(summaries, f)
    return summaries

@targeter
def summarize(targets):
    return [summarize_file(t) for t in targets]

class RequestGetter:
    data = dict()
    def __init__(self, f):
        self.data = read_csv(f)

    def get(self, id):
        return self.data[id]


def save_sort(data, fileroot):
    # Step 1: Create a new Excel Workbook to hold our data.
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet to start with a clean slate.
    rqg = RequestGetter(fileroot)

    # While we're doing that, we'll concurrently make the JSON cleaner
    jsd = dict()


    # Step 2: Loop through each entry in the main dictionary (`data`), which represents a "category."
    for category, subentries in data.items():
        # Create a sheet named after each category.
        sheet = workbook.create_sheet(title=category.replace("/","|"))
        # Step 3: Set headers for this sheet.
        sheet.append(["Topic Name", "Description", "Additional Information"])
        
        # This will save the data for this category
        this_category = dict()

        # Step 4: Populate each row with data from the `requests` list within each subentry.
        for subentry_key, subentry_data in subentries.items():
            # Each subentry contains a list of request IDs under the 'requests' field.
            requests_list = subentry_data.get('requests', [])
            
            # This is the one which changes all the numbers to text for the JSON file
            new_request_list = []

            for request_id in requests_list:
                # Retrieve request details using the get_request function.
                request = rqg.get(request_id)
                new_request_list.append({
                    'Description': request.get('Description', ''),
                    'Additional Information': request.get('Additional Information', '')
                })

                # Append a row with subentry key name and request details.
                sheet.append([
                    subentry_key,
                    request.get('Description', ''),
                    request.get('Additional Information', '')
                ])

            # Now save the new_request_list to this_category
            this_category[subentry_key] = {
                'Description': subentry_data.get('description', ''),
                'Requests': new_request_list
            }

        # Now we can finish the JSON for this category 
        jsd[category] = this_category

        # Step 5: Now create a secondary sheet for each category with " - Topics" appended.
        topics_sheet_name = f"{category} - Topics"
        topics_sheet = workbook.create_sheet(title=topics_sheet_name.replace("/","|"))
        
        # Step 6: Set headers for the " - Topics" sheet.
        topics_sheet.append(["Topic Name", "Description"])
        
        # Step 7: Populate the " - Topics" sheet with each topic's name and description.
        for topic, topic_data in subentries.items():
            topics_sheet.append([
                topic,
                topic_data.get('description', '')
            ])
    
    # Step 8: Save the workbook to an Excel file.
    excel_filename = get_filename(fileroot, "xlsx")
    workbook.save(excel_filename)
    print("Data successfully saved to " + excel_filename)
 
    # Step 9: Save the workbook to a JSON file.
    write_to_json(jsd, fileroot + ".commonalities")
    print(jsd)





def sort_file(f): 
    #   Subject Category Name 1: {
    #       Categorization Name 1: {
    #           desc: Description,
    #           requests: [req ids]
    #       },
    #       ...
    #   },
    #   Subject Category Name 2: ...
    # }
    subjects = dict() # Holds requests corresponding to what topic to what category 
    for i, data in enumerate(SortReader(f)):
        # First, find get the index (The Subject Name)
        index = data['request']['Category']

        # dict.get() will return [] if that entry is null
        subject = subjects.get(index, {})
        dpr("Subject: {}".format(index))
        # Then get a category for the request
        topic: Topic = sort_req(data['summary'], subject)
        if topic is None:
            print("SOMETHING WRONG")
            pass # Something went wrong, I don't care enough to debug 


        # if request_category in current request_categories of subject_category
        if topic.name in subject.keys():
            # Just append the current req id
            subject[topic.name]["requests"].append(i)
        else:
            # Make a list only containing the current req id
            subject[topic.name] = {
                "description": topic.description,
                "requests": [i]
            }

        # Then put the modified subject category back into the big list
        subjects[index] = subject
        save_sort(subjects, f)
        # save_dicts_to_csv(subjects, f + ".commonalities.csv")
    return subjects 

@targeter
def sort(targets):
    return [sort_file(t) for t in targets]

def get_categories(f):
    commonalities = dict()
    with open(get_filename(f + ".commonalities", 'json'), 'r') as fp:
        commonalities = json.load(fp)
    
    return commonalities

def read_sort_file(f):
    commonalities = get_categories(f)    
    requests = read_csv(f)

    for subject in commonalities.keys():
        print("\n\n***Enhancement Category: {}***".format(subject))
        for category in commonalities[subject].keys():
            print("\n**Subject Category: {}**".format(category))
            print("*{}*".format(commonalities[subject][category]['description']))
            for request in commonalities[subject][category]['requests']:
                print(requests[request]['Description'])
                print(requests[request]['Additional Information'], '\n')


@targeter
def read_sort(targets):
    return [read_sort_file(t) for t in targets]

def blurb_file(f):
    # first, we'll just test if all of the data fits into the context length or not
    data = ""
    with open("./data/" + f + ".commonalities.json") as fp:
        data = fp.read()


    res = prompt(PROMPTS["BLURB"], data)
    
    dpr(res)
    with open("./data/" + f + ' Analysis.txt', 'w') as fp:
        fp.write(res)

@targeter
def blurb(targets):
    return [blurb_file(t) for t in targets]


def main():
    parser = argparse.ArgumentParser(
                            prog="CTLS Enhancement Request Summarizer",
                            description="Uses AI to summarize Enhancement Requests for CTLS",
                            epilog="Email AbdulMuqeet.Mohammed@yahoo.com for help")
    subparsers = parser.add_subparsers(dest='command', required=True)

    # Define the targets and actions that require a target (students/teachers)
    actions = ('convert', 'sort', 'read_sort', 'summarize', 'blurb')
    actions_fns = (convert, sort, read_sort, summarize, blurb)

    for action in actions:
        subparser = subparsers.add_parser(action, help=f'{action.capitalize()} target')
        subparser.add_argument('target', default="all", help='Target (sheet) to operate on')


    parser.add_argument('--debug', action='store_true', help='Enable debug mode for verbose output')

    parser.add_argument('--model', choices=MODEL_NICKNAMES.keys(), default='cloud_smart', help='Optionally select model (default: cloud_smart=gpt4o)')


    args = parser.parse_args()

    select_model_from_nickname(args.model)
    
    global DEBUG
    DEBUG = args.debug
   
    fn = actions_fns[actions.index(args.command)]
    fn(args.target)


if __name__ == "__main__":
    main()
